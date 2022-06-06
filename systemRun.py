import cx_Oracle
import random
#connect to DB
ID="system"
PASSWORD="Na123456"
LOCALHOST="localhost/orcl"
DBconnect=cx_Oracle.connect(ID,PASSWORD,LOCALHOST)
cursor=DBconnect.cursor()
#connect to excle
from openpyxl import load_workbook
  #load adjection
EXCLEConnect=load_workbook("./mbti_adjectives.xlsx",data_only=True)
loadSheet=EXCLEConnect['Sheet1']
start='A3'
end='B348'
getCells=loadSheet[start:end]
for row in getCells:
  cursor.execute(" INSERT INTO adjection (adjective,MBTI) VALUES('"+row[0].value+"','"+row[1].value+"')")


  #load name
EXCLEConnect=load_workbook("./name_gender_dataset.xlsx",data_only=True)
loadSheet=EXCLEConnect['Worksheet']
start='A2'
end='B1000'#B147270
getCells=loadSheet[start:end]
for row in getCells:
  cursor.execute(" INSERT INTO name (name,gender) VALUES('"+row[0].value+"','"+row[1].value+"')")
  #load question
EXCLEConnect=load_workbook("./MBTI_question.xlsx",data_only=True)
start='A2'
end='C4'
MBTIType=['M','B','T','I']
for type in MBTIType:
  loadSheet=EXCLEConnect[type]
  getCells=loadSheet[start:end]
  for row in getCells:
    cursor.execute(" INSERT INTO question (type,question,answer1,answer2) VALUES('"+type+"','"+row[0].value+"','"+row[1].value+"','"+row[2].value+"')")
cursor.execute("SELECT * FROM question")
#get user data
userName=input("당신의 이름을 입력해 주세요: ")
while True:
  userGender=input("당신의 성별을 입력해 주세요(M / F): ")
  if(userGender=='M' or userGender=='F'):
    break
  print("잘못된 값을 입력하셨습니다. 다시 입력하세요.")

#have some question of MBTI
  #select sequence
"""
sequence=[0,1,2,3,4,5,6,7,8]
random.shuffle(sequence)
print(sequence)
sequence=[]
count=0
start=1
end=10
while True:
  if(count==9):
    break
  ms=random.randrange(start,end)
  if not(ms in sequence):
    print(ms)
    sequence.append(ms)
    count=count+1
    #make randrange more efficiency
    if(ms==start):
      start=start=1
    if(ms==(end-1)):
      end=end-1
"""

  #let's quest
MBTIType=['M','B','T','I']
answer=[[0,0],[0,0],[0,0],[0,0]]

    #get question from DB
cursor.execute("SELECT * FROM question ORDER BY dbms_random.value")

    #do some quest
for DBsequence in cursor:
  while True:
    ms=int(input(DBsequence[2]))
    if(ms==1 or ms==2):
      for j in range(4):
        if(DBsequence[1]==MBTIType[j]):
          answer[j][ms-1]=answer[j][ms-1]+1
      break
    print("잘못된 값을 입력하셨습니다. 다시 입력하세요.")

#select MBTI
MBTIList=[['I','E'],['N','S'],['T','F'],['P','J']]
userMBTI=""
for i in range(4):
  answerVal=answer[i]
  if(answerVal[0]>answerVal[1]):
    userMBTI=userMBTI+MBTIList[i][0]
  else:
    userMBTI=userMBTI+MBTIList[i][1]

#make english name
  #select adjection
cursor.execute("SELECT adjective FROM adjection WHERE MBTI='"+userMBTI+"' ORDER BY dbms_random.value")
nicknameAdjection=""
for i in cursor:
  nicknameAdjection=i[0]
  break
  #select name
cursor.execute("SELECT name FROM name WHERE gender='"+userGender+"' ORDER BY dbms_random.value")
nicknameName=""
for i in cursor:
  nicknameName=i[0]
  break

#show answer
print(userName+"님의 MBTI는 "+userMBTI+"입니다.")
print(userName+"님의 영어 별명은 "+nicknameAdjection+" "+nicknameName+"입니다.")