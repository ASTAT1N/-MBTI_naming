#for initalize service
#connect to db

#import cx_Oracle
ID="admin"
PASSWORD="Na123456"
LOCALHOST="localhost/orcl"
#DBconnect=cx_Oracle.connect(ID,PASSWORD,LOCALHOST)
#cursor=DBconnect.cursor()
#cursor.execute()
  #make DB of adjection
print("""
CREATE TABLE adjective(
ID int not null
adjective varchar2(100) not null
MBTI char(4) not null
PRIMARY KEY(ID)
);
""")

  #make DB of name
print("""
CREATE TABLE name(
ID int not null
name varchar2(100) not null
gender char(1) not null
PRIMARY KEY(ID)
);
""")

  #make DB of question
print("""
CREATE TABLE question(
ID int not null
type char(1) not null
question varchar(500) not null
answer1 char(1) not null
answer2 char(1) not null
PRIMARY KEY(ID)
);
""")
#connect to excle
import struct
from openpyxl import load_workbook
  #load adjection
EXCLEConnect=load_workbook("./mbti_adjectives.xlsx",data_only=True)
loadSheet=EXCLEConnect['Sheet1']
start='A3'
end='B100'#B348
getCells=loadSheet[start:end]
for row in getCells:
  print(" INSERT INTO adjection (adjective,MBTI) VALUES('"+row[0].value+"','"+row[1].value+"');")

  #load name
EXCLEConnect=load_workbook("./name_gender_dataset.xlsx",data_only=True)
loadSheet=EXCLEConnect['Worksheet']
start='A2'
end='B100'#B147270
getCells=loadSheet[start:end]
for row in getCells:
  print(" INSERT INTO name (name,gender) VALUES('"+row[0].value+"','"+row[1].value+"');")

  #load question
EXCLEConnect=load_workbook("./MBTI_question.xlsx",data_only=True)
start='A2'
end='C4'
MBTIType=['M','B','T','I']
for type in MBTIType:
  loadSheet=EXCLEConnect[type]
  getCells=loadSheet[start:end]
  for row in getCells:
    print(" INSERT INTO question (type,question,answer1,answer2) VALUES('"+type+"','"+row[0].value+"','"+row[1].value+"','"+row[2].value+"');")


